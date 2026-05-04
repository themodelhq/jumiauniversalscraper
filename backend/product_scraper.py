import requests
from bs4 import BeautifulSoup
import os
import re
import json
import time
from urllib.parse import urlparse
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

DEBUG = True

def clean_text(text):
    return re.sub(r'\s+', ' ', text).strip() if text else ""

def log_debug(message):
    if DEBUG:
        print(message)

def create_session():
    """Create a requests session with retry logic and proper headers"""
    session = requests.Session()

    # Configure retry strategy
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET", "OPTIONS"]
    )

    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    # Set default headers that simulate a real browser
    session.headers.update({
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/121.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Cache-Control": "max-age=0"
    })

    return session

def detect_website(url):
    """Detect which website the URL belongs to"""
    domain = urlparse(url).netloc.lower()

    if 'amazon' in domain:
        return 'amazon'
    elif 'noon' in domain:
        return 'noon'
    elif 'gsmarena' in domain:
        return 'gsmarena'
    else:
        return 'generic'

# ============================================================
# AMAZON SCRAPER
# ============================================================
def scrape_amazon_product(url):
    try:
        session = create_session()
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Try to find JSON-LD structured data first
        product_data = extract_json_ld(soup)

        # Product Title
        title = product_data.get('name') or clean_text(soup.select_one("#productTitle").get_text()) if soup.select_one("#productTitle") else ""
        if not title: log_debug(f"⚠️ No Title found for {url}")

        # Brand
        brand = product_data.get('brand') or clean_text(soup.select_one("#bylineInfo").get_text()) if soup.select_one("#bylineInfo") else ""
        if not brand: log_debug(f"⚠️ No Brand found for {url}")

        # Price
        price = product_data.get('price') or ""
        if not price:
            for pid in ["priceblock_ourprice", "priceblock_dealprice", "priceblock_saleprice", "corePrice_feature_div"]:
                tag = soup.select_one(f"span#{pid}")
                if tag:
                    price = clean_text(tag.get_text())
                    break
            if not price:
                price_elem = soup.select_one(".a-price .a-offscreen")
                if price_elem:
                    price = clean_text(price_elem.get_text())
        if not price: log_debug(f"⚠️ No Price found for {url}")

        # Category
        category = product_data.get('category') or ""
        if not category:
            cat1 = " > ".join([clean_text(a.get_text()) for a in soup.select("#wayfinding-breadcrumbs_container li span a")])
            cat2 = " > ".join([clean_text(a.get_text()) for a in soup.select("#wayfinding-breadcrumbs_feature_div ul li span a")])
            category = cat2 if cat2 else cat1
            if not category:
                category = " > ".join([clean_text(a.get_text()) for a in soup.select("div#nav-subnav a, ul.a-unordered-list li a")])
        if not category: log_debug(f"⚠️ No Category found for {url}")

        # Key Features
        features = product_data.get('features') or " | ".join([clean_text(li.get_text()) for li in soup.select("#feature-bullets span.a-list-item")]) or ""
        if not features: log_debug(f"⚠️ No Key Features found for {url}")

        # About this item
        about = ""
        about_header = soup.find("h1", class_="a-size-base-plus a-text-bold", string=re.compile("About this item", re.I))
        if about_header:
            about_list = about_header.find_next("ul")
            if about_list:
                about = " | ".join([clean_text(li.get_text()) for li in about_list.select("li span.a-list-item")])
        if not about: log_debug(f"⚠️ No About This Item found for {url}")

        # Technical & Additional Information
        tech_info = []
        for tid in ["productDetails_techSpec_section_1", "productDetails_detailBullets_sections1"]:
            table = soup.select_one(f"table#{tid}")
            if table:
                for row in table.find_all("tr"):
                    cols = row.find_all(["th", "td"])
                    if len(cols) >= 2:
                        key = clean_text(cols[0].get_text())
                        val = clean_text(cols[1].get_text())
                        tech_info.append(f"{key}: {val}")
        if not tech_info:
            for header in soup.find_all(["h2", "h1"], string=re.compile("Product information|Technical Details|Additional Information", re.I)):
                table = header.find_next("table")
                if table:
                    for row in table.find_all("tr"):
                        cols = row.find_all(["th", "td"])
                        if len(cols) >= 2:
                            key = clean_text(cols[0].get_text())
                            val = clean_text(cols[1].get_text())
                            tech_info.append(f"{key}: {val}")
        tech_info_text = " | ".join(tech_info)
        if not tech_info_text: log_debug(f"⚠️ No Technical/Additional Info found for {url}")

        # Product Description
        desc = product_data.get('description') or ""
        if not desc:
            if soup.select_one("#productDescription"):
                desc += " " + clean_text(soup.select_one("#productDescription").get_text())
            if soup.select_one("#aplus"):
                desc += " " + clean_text(soup.select_one("#aplus").get_text())
            h2_desc = soup.find("h2", class_="default", string=re.compile("Product description", re.I))
            if h2_desc:
                desc_section = h2_desc.find_next()
                if desc_section:
                    desc += " " + clean_text(desc_section.get_text())
        if not desc.strip(): log_debug(f"⚠️ No Product Description found for {url}")

        # Product Images
        images = product_data.get('images', []) or []

        img_wrapper = soup.select_one("#imgTagWrapperId img")
        if img_wrapper:
            for attr in ("data-old-hires", "src"):
                val = img_wrapper.get(attr)
                if val and val not in images:
                    images.append(val)

        landing_img = soup.select_one("#landingImage")
        if landing_img:
            for attr in ("data-old-hires", "src"):
                val = landing_img.get(attr)
                if val and val not in images:
                    images.append(val)
            if landing_img.has_attr("data-a-dynamic-image"):
                try:
                    dyn_imgs = json.loads(landing_img["data-a-dynamic-image"])
                    for src in dyn_imgs.keys():
                        if src not in images:
                            images.append(src)
                except:
                    pass

        script = soup.find("script", string=re.compile(r'ImageBlockATF|colorImages'))
        if script:
            text = script.string
            match = re.search(r'(?P<json>\{.*"colorImages".*?\})', text, re.DOTALL)
            if match:
                try:
                    data = json.loads(match.group("json"))
                    for img in data.get("colorImages", {}).get("initial", []):
                        hi = img.get("hiRes") or img.get("large") or img.get("thumb")
                        if hi and hi not in images:
                            images.append(hi)
                except Exception:
                    pass

        for img in soup.find_all("img", src=True):
            src = img.get("data-old-hires") or img.get("src")
            if src and src not in images:
                images.append(src)

        if not images: log_debug(f"⚠️ No Images found for {url}")

        return {
            "URL": url,
            "Website": "Amazon",
            "Product Name": title,
            "Brand": brand,
            "Price": price,
            "Currency": product_data.get('currency') or "USD",
            "Category": category,
            "Key Features": features,
            "About This Item": about,
            "Tech & Additional Info": tech_info_text,
            "Description": desc.strip(),
            "Rating": product_data.get('rating') or "",
            "Reviews Count": product_data.get('review_count') or "",
            "SKU": product_data.get('sku') or "",
            "Availability": product_data.get('availability') or "",
            "Images": " | ".join(images) if isinstance(images, list) else images
        }
    except Exception as e:
        log_debug(f"❌ Error scraping {url}: {e}")
        return {"URL": url, "Website": "Amazon", "Error": str(e)}

# ============================================================
# NOON SCRAPER
# ============================================================
def scrape_noon_product(url):
    try:
        session = create_session()

        # Noon may require specific headers or cookies
        session.headers.update({
            "Accept-Language": "en-US,en;q=0.9,ar;q=0.8",
            "Referer": "https://www.noon.com/"
        })

        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Try JSON-LD first
        product_data = extract_json_ld(soup)

        # Product Title
        title = product_data.get('name') or ""
        if not title:
            # Try multiple selectors for Noon
            for selector in ["h1.productTitle", "h1[itemprop='name']", ".product-title h1",
                            "[data-qa='product-title']", "h1"]:
                elem = soup.select_one(selector)
                if elem:
                    text = clean_text(elem.get_text())
                    if text and len(text) > 5:
                        title = text
                        break
        if not title: log_debug(f"⚠️ No Title found for {url}")

        # Brand
        brand = product_data.get('brand') or ""
        if not brand:
            for selector in ["a.product-brand", ".productBrand a", "[data-qa='product-brand']",
                            ".brand-name", "[itemprop='brand']", ".seller-name"]:
                elem = soup.select_one(selector)
                if elem:
                    brand = clean_text(elem.get_text())
                    break
        if not brand: log_debug(f"⚠️ No Brand found for {url}")

        # Price
        price = product_data.get('price') or ""
        currency = product_data.get('currency') or "AED"
        if not price:
            for selector in [".price", ".product-price", "[data-qa='product-price']",
                            ".current-price", "[itemprop='price']", ".selling-price"]:
                elem = soup.select_one(selector)
                if elem:
                    price = clean_text(elem.get_text())
                    break
        if not price:
            # Try to find any price-like element
            price_pattern = soup.find(string=re.compile(r'[\d,]+\s*(AED|SAR|EGP)'))
            if price_pattern:
                price = clean_text(price_pattern)
        if not price: log_debug(f"⚠️ No Price found for {url}")

        # Category
        category = product_data.get('category') or ""
        if not category:
            cats = []
            for selector in [".breadcrumb a", ".product-breadcrumb a", "[data-qa='breadcrumb'] a",
                            ".nav-breadcrumb a", "[itemprop='breadcrumb'] a"]:
                for a in soup.select(selector):
                    cat_text = clean_text(a.get_text())
                    if cat_text and cat_text.lower() not in ['home', 'products', 'product', 'noon']:
                        cats.append(cat_text)
            category = " > ".join(cats)
        if not category: log_debug(f"⚠️ No Category found for {url}")

        # Description
        description = product_data.get('description') or ""
        if not description:
            for selector in [".product-description", ".description", "[data-qa='product-description']",
                            "[itemprop='description']", ".product-details"]:
                elem = soup.select_one(selector)
                if elem:
                    description = clean_text(elem.get_text())
                    break
        if not description:
            # Try finding any description element
            for selector in ["[data-qa='product-details']", ".details-section"]:
                elem = soup.select_one(selector)
                if elem:
                    description = clean_text(elem.get_text())
                    break
        if not description: log_debug(f"⚠️ No Description found for {url}")

        # Images
        images = product_data.get('images', []) or []
        if not images:
            for selector in [".product-images img", ".gallery img", "[data-qa='product-image']",
                            "[itemprop='image']", ".main-image img"]:
                for img in soup.select(selector):
                    src = img.get('src') or img.get('data-src') or img.get('data-qa')
                    if src and src not in images and 'placeholder' not in src.lower():
                        images.append(src)

        # Try JSON data in script tags
        if not images:
            scripts = soup.find_all("script", type="application/ld+json")
            for script in scripts:
                try:
                    data = json.loads(script.string)
                    if isinstance(data, dict) and 'image' in data:
                        img_list = data['image'] if isinstance(data['image'], list) else [data['image']]
                        for img in img_list:
                            if img and img not in images:
                                images.append(img)
                except:
                    pass

        # Also check for Next.js data
        if not images:
            for script in soup.find_all("script"):
                if script.string and ('image' in script.string.lower() or 'gallery' in script.string.lower()):
                    # Try to find JSON-like image data
                    img_matches = re.findall(r'(?:https?://[^"\'>\s]+\.(?:jpg|jpeg|png|webp))', script.string)
                    for img in img_matches:
                        if img and img not in images:
                            images.append(img)

        if not images: log_debug(f"⚠️ No Images found for {url}")

        # Rating
        rating = product_data.get('rating') or ""
        if not rating:
            for selector in [".rating", "[itemprop='ratingValue']", ".product-rating",
                            "[data-qa='product-rating']"]:
                elem = soup.select_one(selector)
                if elem:
                    rating = clean_text(elem.get_text())
                    break

        # Reviews Count
        reviews_count = product_data.get('review_count') or ""
        if not reviews_count:
            for selector in [".reviews-count", "[itemprop='reviewCount']", ".review-count",
                            "[data-qa='reviews-count']"]:
                elem = soup.select_one(selector)
                if elem:
                    reviews_count = clean_text(elem.get_text())
                    break

        # SKU
        sku = product_data.get('sku') or ""
        if not sku:
            for selector in ["[data-qa='product-sku']", ".sku", ".product-sku", "[itemprop='sku']"]:
                elem = soup.select_one(selector)
                if elem:
                    sku = clean_text(elem.get_text())
                    break

        # Availability
        availability = product_data.get('availability') or ""
        if not availability:
            for selector in [".availability", "[itemprop='availability']", ".stock-status",
                            ".delivery-info", "[data-qa='availability']"]:
                elem = soup.select_one(selector)
                if elem:
                    availability = clean_text(elem.get_text())
                    break

        return {
            "URL": url,
            "Website": "Noon",
            "Product Name": title,
            "Brand": brand,
            "Price": price,
            "Currency": currency,
            "Category": category,
            "Key Features": "",
            "About This Item": "",
            "Tech & Additional Info": "",
            "Description": description,
            "Rating": rating,
            "Reviews Count": reviews_count,
            "SKU": sku,
            "Availability": availability,
            "Images": " | ".join(images) if isinstance(images, list) else images
        }
    except Exception as e:
        log_debug(f"❌ Error scraping {url}: {e}")
        return {"URL": url, "Website": "Noon", "Error": str(e)}

# ============================================================
# GSM ARENA SCRAPER
# ============================================================
def scrape_gsmarena_product(url):
    try:
        session = create_session()
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Try JSON-LD first
        product_data = extract_json_ld(soup)

        # Product Title (phone name)
        title = product_data.get('name') or ""
        if not title:
            for selector in ["h1.article-title", "h1.specs-phone-name-title", ".help-box h1",
                            "h1", ".product-name"]:
                elem = soup.select_one(selector)
                if elem:
                    text = clean_text(elem.get_text())
                    if text and len(text) > 3:
                        title = text
                        break
        if not title: log_debug(f"⚠️ No Title found for {url}")

        # Brand
        brand = product_data.get('brand') or ""
        if not brand:
            # Extract brand from URL or page
            url_path = urlparse(url).path
            brand_patterns = {
                'samsung': 'Samsung', 'apple': 'Apple', 'xiaomi': 'Xiaomi',
                'huawei': 'Huawei', 'oneplus': 'OnePlus', 'google': 'Google',
                'oppo': 'Oppo', 'vivo': 'Vivo', 'realme': 'Realme', 'motorola': 'Motorola',
                'nokia': 'Nokia', 'sony': 'Sony', 'asus': 'Asus', 'Lenovo': 'Lenovo',
                'honor': 'Honor', 'tecno': 'Tecno', 'infinix': 'Infinix'
            }
            for pattern, brand_name in brand_patterns.items():
                if pattern in url_path.lower():
                    brand = brand_name
                    break

        # Price
        price = product_data.get('price') or ""
        currency = product_data.get('currency') or "USD"
        if not price:
            for selector in [".help-box .price", ".price", "[itemprop='price']", ".specs-phone-price"]:
                elem = soup.select_one(selector)
                if elem:
                    price = clean_text(elem.get_text())
                    break
        if not price:
            # Try finding price text patterns
            price_elem = soup.find(string=re.compile(r'^\s*[\d,]+\s*(USD|EUR|GBP)?', re.I))
            if price_elem:
                price = clean_text(price_elem)
        if not price: log_debug(f"⚠️ No Price found for {url}")

        # Category
        category = product_data.get('category') or "Mobile Phone"
        if not category:
            if soup.select_one(".nav-subbrand a"):
                category = clean_text(soup.select_one(".nav-subbrand a").get_text())

        # Specification tables - extract all specs
        specs = {}

        # Main specs table
        for table in soup.select("table.specs-list"):
            for row in table.select("tr"):
                cells = row.find_all(["th", "td"])
                if len(cells) >= 2:
                    key = clean_text(cells[0].get_text())
                    val = clean_text(cells[1].get_text())
                    if key and val:
                        specs[key] = val

        # Alternate specs structure - div based
        for div in soup.select(".specs-list, .specs-border"):
            for row in div.select(".tr"):
                key_elem = row.select_one(".ttl")
                val_elem = row.select_one(".nfo")
                if key_elem and val_elem:
                    key = clean_text(key_elem.get_text())
                    val = clean_text(val_elem.get_text())
                    if key and val:
                        specs[key] = val

        # Format specs as text
        tech_info = " | ".join([f"{k}: {v}" for k, v in specs.items()])
        if not tech_info: log_debug(f"⚠️ No Technical Info found for {url}")

        # Description
        description = product_data.get('description') or ""
        if not description:
            for selector in [".article-body", ".review-body", "[itemprop='description']",
                            ".article-text"]:
                elem = soup.select_one(selector)
                if elem:
                    description = clean_text(elem.get_text())
                    break
        if not description:
            intro = soup.select_one(".article-info")
            if intro:
                description = clean_text(intro.get_text())
        if not description:
            # Get meta description as fallback
            meta_desc = soup.select_one("meta[name='description']")
            if meta_desc:
                description = meta_desc.get('content', '')
        if not description: log_debug(f"⚠️ No Description found for {url}")

        # Images
        images = product_data.get('images', []) or []
        if not images:
            for selector in ["[itemprop='image']", ".specs-photo-main img", ".article-img img",
                            ".gallery-image img", ".phone-image img"]:
                elem = soup.select_one(selector)
                if elem:
                    src = elem.get('src') or elem.get('data-src') or elem.get('data-lazy')
                    if src and src not in images and 'placeholder' not in src.lower():
                        images.append(src)

            # Also check for multiple images in gallery
            for img in soup.select("img"):
                src = img.get('src') or img.get('data-src')
                if src and src not in images and 'placeholder' not in src.lower():
                    if any(x in src.lower() for x in ['phone', 'gallery', 'specs', 'device']):
                        images.append(src)

        if not images: log_debug(f"⚠️ No Images found for {url}")

        return {
            "URL": url,
            "Website": "GSMArena",
            "Product Name": title,
            "Brand": brand,
            "Price": price,
            "Currency": currency,
            "Category": category,
            "Key Features": "",
            "About This Item": "",
            "Tech & Additional Info": tech_info,
            "Description": description,
            "Rating": product_data.get('rating') or "",
            "Reviews Count": product_data.get('review_count') or "",
            "SKU": product_data.get('sku') or "",
            "Availability": "",
            "Images": " | ".join(images) if isinstance(images, list) else images
        }
    except Exception as e:
        log_debug(f"❌ Error scraping {url}: {e}")
        return {"URL": url, "Website": "GSMArena", "Error": str(e)}

# ============================================================
# GENERIC/FLEXIBLE SCRAPER (For any website)
# ============================================================
def scrape_generic_product(url):
    """Generic scraper that attempts to extract product data from any website"""
    try:
        session = create_session()
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Try JSON-LD structured data first (most reliable for any site)
        product_data = extract_json_ld(soup)

        # Product Title
        title = product_data.get('name') or ""
        if not title:
            selectors = [
                "h1[itemprop='name']",
                "[itemprop='name']",
                "h1.product-title",
                "h1.product-name",
                "h1.title",
                ".product-title h1",
                ".product-name h1",
                ".product h1",
                "h1",
                ".title h1"
            ]
            for selector in selectors:
                elem = soup.select_one(selector)
                if elem:
                    text = clean_text(elem.get_text())
                    if text and len(text) > 3:
                        title = text
                        break
        if not title: log_debug(f"⚠️ No Title found for {url}")

        # Brand
        brand = product_data.get('brand') or ""
        if not brand:
            selectors = [
                "[itemprop='brand']",
                ".product-brand",
                ".brand",
                ".product-brand a",
                "[data-brand]"
            ]
            for selector in selectors:
                elem = soup.select_one(selector)
                if elem:
                    brand = clean_text(elem.get_text())
                    break
        if not brand: log_debug(f"⚠️ No Brand found for {url}")

        # Price
        price = product_data.get('price') or ""
        currency = product_data.get('currency') or ""
        if not price:
            selectors = [
                "[itemprop='price']",
                ".price",
                ".product-price",
                ".current-price",
                ".sale-price",
                ".regular-price",
                "[data-price]",
                ".amount"
            ]
            for selector in selectors:
                elem = soup.select_one(selector)
                if elem:
                    price = clean_text(elem.get_text())
                    if not currency:
                        currency_match = re.search(r'([₹$€£¥])', price)
                        if currency_match:
                            currency_map = {'₹': 'INR', '$': 'USD', '€': 'EUR', '£': 'GBP', '¥': 'CNY'}
                            currency = currency_map.get(currency_match.group(1), '')
                    break
        if not price: log_debug(f"⚠️ No Price found for {url}")

        # Category
        category = product_data.get('category') or ""
        if not category:
            selectors = [".breadcrumb", ".breadcrumbs", ".nav-crumb", "[itemprop='breadcrumb']"]
            for selector in selectors:
                cats = []
                for a in soup.select(f"{selector} a"):
                    cat_text = clean_text(a.get_text())
                    if cat_text and cat_text.lower() not in ['home', 'products', 'product']:
                        cats.append(cat_text)
                if cats:
                    category = " > ".join(cats)
                    break
        if not category: log_debug(f"⚠️ No Category found for {url}")

        # Description
        description = product_data.get('description') or ""
        if not description:
            selectors = [
                "[itemprop='description']",
                ".product-description",
                ".description",
                ".product-details",
                "#description",
                ".description-content"
            ]
            for selector in selectors:
                elem = soup.select_one(selector)
                if elem:
                    description = clean_text(elem.get_text())
                    break
        if not description: log_debug(f"⚠️ No Description found for {url}")

        # Images
        images = product_data.get('images', []) or []
        if not images:
            selectors = [
                "[itemprop='image']",
                ".product-image img",
                ".product-gallery img",
                ".gallery img",
                ".product-img img",
                ".main-image img",
                "#product-image",
                "#main-image"
            ]
            for selector in selectors:
                for img in soup.select(selector):
                    src = img.get('src') or img.get('data-src') or img.get('data-lazy')
                    if src and src not in images:
                        images.append(src)
                    srcset = img.get('srcset')
                    if srcset:
                        first_src = srcset.split(',')[0].split()[0]
                        if first_src and first_src not in images:
                            images.append(first_src)
        if not images: log_debug(f"⚠️ No Images found for {url}")

        # Rating
        rating = product_data.get('rating') or ""
        if not rating:
            for selector in ["[itemprop='ratingValue']", ".rating", ".stars", "[data-rating]"]:
                elem = soup.select_one(selector)
                if elem:
                    rating = clean_text(elem.get_text())
                    break

        # Reviews Count
        reviews_count = product_data.get('review_count') or ""
        if not reviews_count:
            for selector in ["[itemprop='reviewCount']", ".review-count", ".reviews-count"]:
                elem = soup.select_one(selector)
                if elem:
                    reviews_count = clean_text(elem.get_text())
                    break

        # SKU
        sku = product_data.get('sku') or ""
        if not sku:
            for selector in ["[itemprop='sku']", ".sku", ".product-code", "[data-sku]"]:
                elem = soup.select_one(selector)
                if elem:
                    sku = clean_text(elem.get_text())
                    break

        # Availability
        availability = product_data.get('availability') or ""
        if not availability:
            for selector in ["[itemprop='availability']", ".availability", ".stock", ".product-stock"]:
                elem = soup.select_one(selector)
                if elem:
                    availability = clean_text(elem.get_text())
                    break

        return {
            "URL": url,
            "Website": "Generic",
            "Product Name": title,
            "Brand": brand,
            "Price": price,
            "Currency": currency,
            "Category": category,
            "Key Features": "",
            "About This Item": "",
            "Tech & Additional Info": "",
            "Description": description,
            "Rating": rating,
            "Reviews Count": reviews_count,
            "SKU": sku,
            "Availability": availability,
            "Images": " | ".join(images) if isinstance(images, list) else images
        }
    except Exception as e:
        log_debug(f"❌ Error scraping {url}: {e}")
        return {"URL": url, "Website": "Generic", "Error": str(e)}

# ============================================================
# HELPER FUNCTIONS
# ============================================================
def extract_json_ld(soup):
    """Extract product data from JSON-LD structured data"""
    data = {}

    scripts = soup.find_all("script", type="application/ld+json")

    for script in scripts:
        try:
            content = script.string
            if content:
                json_data = json.loads(content)

                if isinstance(json_data, dict):
                    if json_data.get('@type') == 'Product' or 'Product' in str(json_data.get('@type', '')):
                        data = extract_product_from_json_ld(json_data)
                        break
                    if '@graph' in json_data:
                        for item in json_data['@graph']:
                            if item.get('@type') == 'Product':
                                data = extract_product_from_json_ld(item)
                                break
                elif isinstance(json_data, list):
                    for item in json_data:
                        if isinstance(item, dict) and (item.get('@type') == 'Product' or 'Product' in str(item.get('@type', ''))):
                            data = extract_product_from_json_ld(item)
                            break
        except (json.JSONDecodeError, TypeError) as e:
            log_debug(f"JSON-LD parse error: {e}")
            continue

    return data

def extract_product_from_json_ld(product):
    """Extract relevant fields from JSON-LD Product object"""
    data = {}

    if 'name' in product:
        data['name'] = product['name']

    if 'brand' in product:
        if isinstance(product['brand'], str):
            data['brand'] = product['brand']
        elif isinstance(product['brand'], dict) and 'name' in product['brand']:
            data['brand'] = product['brand']['name']

    if 'description' in product:
        data['description'] = product['description']

    if 'sku' in product:
        data['sku'] = product['sku']

    if 'image' in product:
        images = product['image']
        if isinstance(images, str):
            data['images'] = [images]
        elif isinstance(images, list):
            data['images'] = images
        elif isinstance(images, dict) and 'url' in images:
            data['images'] = [images['url']]

    if 'offers' in product:
        offers = product['offers']
        if isinstance(offers, dict):
            if 'price' in offers:
                data['price'] = offers['price']
            if 'priceCurrency' in offers:
                data['currency'] = offers['priceCurrency']
            if 'availability' in offers:
                data['availability'] = offers['availability']
        elif isinstance(offers, list) and offers:
            first = offers[0]
            if isinstance(first, dict):
                if 'price' in first:
                    data['price'] = first['price']
                if 'priceCurrency' in first:
                    data['currency'] = first['priceCurrency']
                if 'availability' in first:
                    data['availability'] = first['availability']

    if 'aggregateRating' in product:
        rating_data = product['aggregateRating']
        if 'ratingValue' in rating_data:
            data['rating'] = rating_data['ratingValue']
        if 'reviewCount' in rating_data:
            data['review_count'] = rating_data['reviewCount']

    if 'category' in product:
        data['category'] = product['category']

    return data

def scrape_product(url):
    """Main dispatcher - routes to appropriate scraper based on URL"""
    website_type = detect_website(url)

    if website_type == 'amazon':
        result = scrape_amazon_product(url)
    elif website_type == 'noon':
        result = scrape_noon_product(url)
    elif website_type == 'gsmarena':
        result = scrape_gsmarena_product(url)
    else:
        result = scrape_generic_product(url)

    # If specialized scraper fails or returns minimal data, try generic as fallback
    if "Error" in result or not result.get("Product Name"):
        log_debug(f"Falling back to generic scraper for {url}")
        fallback = scrape_generic_product(url)
        # Merge results, keeping any successful data from specialized scraper
        for key in fallback:
            if key not in result or not result[key]:
                result[key] = fallback[key]

    return result

# ============================================================
# MAIN RUNNER
# ============================================================
def run_scraper(infile, outfile, delay=2):
    """Process input file and scrape all URLs — CLI use only."""
    import csv as _csv
    ext = os.path.splitext(infile)[1].lower()
    urls = []
    if ext == ".csv":
        with open(infile, newline="", encoding="utf-8-sig") as f:
            reader = _csv.DictReader(f)
            col = next((c for c in (reader.fieldnames or []) if "url" in c.lower()), None)
            if not col:
                raise ValueError("No URL column found in input file.")
            urls = [row[col].strip() for row in reader if row.get(col,"").strip().startswith(("http://","https://"))]
    elif ext in [".xls", ".xlsx"]:
        import openpyxl
        wb = openpyxl.load_workbook(infile)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in ws[1]]
        col_idx = next((i for i, h in enumerate(headers) if "url" in h.lower()), 0)
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = str(row[col_idx] or "").strip()
            if val.startswith(("http://","https://")):
                urls.append(val)
    else:
        raise ValueError("File must be a CSV or Excel (.xls/.xlsx)")

    results = []
    for idx, url_str in enumerate(urls):
        print(f"🔎 [{idx+1}/{len(urls)}] Scraping: {url_str}")
        result = scrape_product(url_str)
        results.append(result)
        if idx < len(urls) - 1:
            time.sleep(delay)

    fieldnames = ["URL","Website","Product Name","Brand","Price","Currency",
                  "Category","Key Features","About This Item","Tech & Additional Info",
                  "Description","Rating","Reviews Count","SKU","Availability","Images"]
    with open(outfile, "w", newline="", encoding="utf-8-sig") as f:
        w = _csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in results:
            row = {k: r.get(k, "") for k in fieldnames}
            w.writerow(row)

    success = sum(1 for r in results if not r.get("Error"))
    print(f"\n✅ Done! {success} successful, {len(results)-success} errors → {outfile}")
    return results


if __name__ == "__main__":
    try:
        from tkinter import Tk, filedialog
        Tk().withdraw()
        infile = filedialog.askopenfilename(
            title="Select a file with product URLs",
            filetypes=[("Excel/CSV files", "*.csv *.xls *.xlsx")]
        )
        if infile:
            outfile = filedialog.asksaveasfilename(
                title="Save output file as",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv")]
            )
            if outfile:
                run_scraper(infile, outfile)
            else:
                print("❌ No output file selected.")
        else:
            print("❌ No input file selected.")
    except ImportError:
        print("Run from command line: python product_scraper.py")
