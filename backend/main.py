"""
Universal Product Scraper — FastAPI Backend
Deploy on Render (render.com) as a Python web service.
No pandas dependency — pure Python + openpyxl only.
"""

from fastapi import FastAPI, BackgroundTasks, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import List
import uuid, csv, time, shutil, re, io
from pathlib import Path

from product_scraper import scrape_product

# Brand website search pipeline (from scraper_gui.py)
try:
    from brand_scraper import search_brand_website, _ensure_hd_images, _extract_gtin
    BRAND_SEARCH_AVAILABLE = True
except Exception:
    BRAND_SEARCH_AVAILABLE = False

# Category page scraper (from scraper_gui.py)
try:
    from category_scraper import scrape_category_page
    CATEGORY_SCRAPER_AVAILABLE = True
except Exception:
    CATEGORY_SCRAPER_AVAILABLE = False

app = FastAPI(title="Universal Product Scraper API", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

JOBS: dict = {}
RESULTS_DIR = Path("results")
RESULTS_DIR.mkdir(exist_ok=True)
TPL = Path(__file__).parent


class ScrapeURLRequest(BaseModel):
    url: str
    format: str = "bob"

class BatchScrapeRequest(BaseModel):
    urls: List[str]
    format: str = "bob"
    delay: float = 1.5

class BrandQueryRequest(BaseModel):
    query: str
    format: str = "bob"

class BatchQueryRequest(BaseModel):
    queries: List[str]
    format: str = "bob"
    delay: float = 2.0

class CategoryScrapeRequest(BaseModel):
    url: str
    max_products: int = 50
    format: str = "bob"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _colour(r):
    n = r.get("Product Name", "")
    for c in ["Black","White","Silver","Gold","Blue","Red","Green","Grey","Gray","Rose Gold"]:
        if c.lower() in n.lower(): return c
    return r.get("Colour", "")

def _imgs(r):
    return [u.strip() for u in re.split(r"[|,]", r.get("Images","")) if u.strip()]

def _job(jid, total=0):
    JOBS[jid] = {"id":jid,"status":"pending","progress":0,"message":"Starting…",
                 "total":total,"completed":0,"results":[],"download_url":None,"error":None}
    return JOBS[jid]

def _read_urls_from_bytes(content: bytes, filename: str) -> list:
    """Read URLs from CSV or Excel bytes without pandas."""
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return []
        col = next((c for c in reader.fieldnames or [] if "url" in c.lower()),
                   list(rows[0].keys())[0] if rows else None)
        if not col:
            return []
        return [str(row[col]).strip() for row in rows
                if row.get(col, "").strip().startswith(("http://", "https://"))]
    else:
        # Excel via openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        url_col_idx = next((i for i, h in enumerate(headers) if "url" in h.lower()), 0)
        urls = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            val = str(row[url_col_idx] or "").strip()
            if val.startswith(("http://", "https://")):
                urls.append(val)
        return urls

def _save_raw_csv(results, path):
    if not results:
        path.write_text("", encoding="utf-8-sig")
        return
    fields = list(results[0].keys())
    # Ensure all fields are covered
    all_fields = []
    seen = set()
    for r in results:
        for k in r:
            if k not in seen:
                seen.add(k); all_fields.append(k)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=all_fields, extrasaction="ignore")
        w.writeheader(); w.writerows(results)

def _save_bob(results, path):
    tpl = TPL / "BOBTemplate.csv"
    with open(tpl, newline="", encoding="utf-8-sig") as f:
        fields = list(csv.DictReader(f).fieldnames or [])
    rows = []
    for r in results:
        row = {f:"" for f in fields}
        imgs = _imgs(r)
        row.update({"brand":r.get("Brand",""),"name":r.get("Product Name",""),
            "model":r.get("SKU",""),"price":r.get("Price",""),
            "color_family":_colour(r),"product_weight":r.get("Weight",""),
            "product_warranty":r.get("Warranty",""),"product_measures":r.get("Dimensions",""),
            "Category":r.get("Category",""),"short_description":r.get("Key Features",""),
            "Description":r.get("Description","") or r.get("About This Item",""),
            "Product Attribute":r.get("Tech & Additional Info",""),
            "Image file names":", ".join(imgs),
            "Has long desription":"Yes" if r.get("Description") else "No"})
        for k in ("gtin_ barcode","gtin_barcode","GTIN"):
            if k in fields: row[k] = r.get("GTIN","")
        for i,col in enumerate(["Image1","Image2","Image3","Image4","Image5","Image6","Image7","Image8"]):
            if col in fields and i < len(imgs): row[col] = imgs[i]
        rows.append(row)
    with open(path,"w",newline="",encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields); w.writeheader(); w.writerows(rows)

def _save_vendor(results, path):
    import openpyxl; from openpyxl.styles import Alignment
    shutil.copy(TPL/"VendorCenterTemplate.xlsx", path)
    wb = openpyxl.load_workbook(path); ws = wb["Upload Template"]
    hdr = {str(c.value).strip(): c.column for c in ws[1] if c.value}
    def _s(rn, col, val):
        c = hdr.get(col)
        if c and val:
            cell = ws.cell(row=rn, column=c, value=str(val))
            if len(str(val))>100: cell.alignment = Alignment(wrap_text=True)
    for i,r in enumerate(results):
        rn=2+i; imgs=_imgs(r)
        _s(rn,"Name",r.get("Product Name","")); _s(rn,"Brand",r.get("Brand",""))
        _s(rn,"Description",r.get("Description","") or r.get("About This Item",""))
        _s(rn,"short_description",r.get("Key Features",""))
        _s(rn,"PrimaryCategory",r.get("Category","")); _s(rn,"GTIN_Barcode",r.get("GTIN",""))
        _s(rn,"Price_NGN",r.get("Price","")); _s(rn,"model",r.get("SKU",""))
        _s(rn,"product_warranty",r.get("Warranty","")); _s(rn,"product_weight",r.get("Weight",""))
        _s(rn,"product_measures",r.get("Dimensions","")); _s(rn,"color",_colour(r))
        _s(rn,"note",r.get("Tech & Additional Info",""))
        for j,col in enumerate(["MainImage","Image2","Image3","Image4","Image5","Image6","Image7","Image8"]):
            if j<len(imgs): _s(rn,col,imgs[j])
    wb.save(path)

def _save(results, fmt, jid):
    if fmt=="bob":     p=RESULTS_DIR/f"{jid}_BOB.csv";     _save_bob(results,p)
    elif fmt=="vendor":p=RESULTS_DIR/f"{jid}_Vendor.xlsx"; _save_vendor(results,p)
    else:              p=RESULTS_DIR/f"{jid}_raw.csv";      _save_raw_csv(results,p)
    return p.name

import concurrent.futures as _cf

def _scrape_safe(url, timeout_sec=45):
    """Scrape a single URL with a hard timeout so hung requests don't freeze the batch."""
    with _cf.ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(scrape_product, url)
        try:
            return fut.result(timeout=timeout_sec)
        except _cf.TimeoutError:
            return {"URL": url, "Error": f"Timed out after {timeout_sec}s"}
        except Exception as e:
            return {"URL": url, "Error": str(e)}


def _run(jid, urls, fmt, delay):
    j = JOBS[jid]; j["status"] = "running"; results = []
    for i, url in enumerate(urls):
        j["message"] = f"Scraping {i+1}/{len(urls)}: {url[:55]}\u2026"
        j["progress"] = int(i / len(urls) * 90)
        r = _scrape_safe(url)
        results.append(r); j["results"].append(r)
        j["completed"] = i + 1
        if i < len(urls) - 1:
            time.sleep(delay)
    try:
        fn = _save(results, fmt, jid); j["download_url"] = f"/download/{fn}"
    except Exception as e:
        j["error"] = str(e)
    j.update({"status": "done", "progress": 100,
               "message": f"Done \u2014 {len(results)} products scraped"})


@app.get("/")
def root(): return {"status":"Universal Product Scraper API v1.0"}
@app.get("/health")
def health(): return {"status":"ok"}

@app.post("/scrape/url")
def scrape_url(req: ScrapeURLRequest):
    if not req.url.startswith(("http://","https://")): raise HTTPException(400,"Invalid URL")
    try:
        result=scrape_product(req.url); jid=uuid.uuid4().hex[:8]
        fn=_save([result],req.format,jid)
        return {"success":True,"result":result,"download":f"/download/{fn}"}
    except Exception as e: raise HTTPException(500,str(e))

@app.post("/scrape/batch")
def scrape_batch(req: BatchScrapeRequest, background_tasks: BackgroundTasks):
    urls=[u for u in req.urls if u.strip().startswith(("http://","https://"))]
    if not urls: raise HTTPException(400,"No valid URLs")
    jid=uuid.uuid4().hex[:8]; _job(jid,len(urls))
    background_tasks.add_task(_run,jid,urls,req.format,req.delay)
    return {"job_id":jid,"total":len(urls)}

@app.post("/scrape/file")
async def scrape_file(background_tasks: BackgroundTasks,
                      file: UploadFile=File(...), format: str=Form("bob"), delay: float=Form(1.5)):
    content = await file.read()
    try:
        urls = _read_urls_from_bytes(content, file.filename)
    except Exception as e:
        raise HTTPException(400, f"Cannot parse: {e}")
    if not urls: raise HTTPException(400,"No valid URLs in file")
    jid=uuid.uuid4().hex[:8]; _job(jid,len(urls))
    background_tasks.add_task(_run,jid,urls,format,delay)
    return {"job_id":jid,"total":len(urls)}

@app.get("/job/{job_id}")
def get_job(job_id: str):
    j=JOBS.get(job_id)
    if not j: raise HTTPException(404,"Job not found")
    return j


# ── Brand query (single) ──────────────────────────────────────────────────────
@app.post("/scrape/query")
def scrape_query(req: BrandQueryRequest):
    """
    Search brand official websites for a product by query string.
    Format: 'Brand Model Colour Storage ModelCode'
    e.g. 'Samsung Galaxy A06 Black 128GB SM-A065FZKHAFB'
    """
    if not req.query.strip():
        raise HTTPException(400, "Query cannot be empty")
    if not BRAND_SEARCH_AVAILABLE:
        raise HTTPException(503, "Brand search module not available")
    try:
        result = search_brand_website(req.query.strip())
        result = _ensure_hd_images(result)
        if not result.get("GTIN"):
            result["GTIN"] = _extract_gtin(result)
        jid = uuid.uuid4().hex[:8]
        fn = _save([result], req.format, jid)
        return {"success": True, "result": result, "download": f"/download/{fn}"}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Brand query batch ─────────────────────────────────────────────────────────
@app.post("/scrape/batch-query")
def scrape_batch_query(req: BatchQueryRequest, background_tasks: BackgroundTasks):
    queries = [q.strip() for q in req.queries if q.strip()]
    if not queries:
        raise HTTPException(400, "No queries provided")
    if not BRAND_SEARCH_AVAILABLE:
        raise HTTPException(503, "Brand search module not available")
    jid = uuid.uuid4().hex[:8]
    _job(jid, len(queries))
    background_tasks.add_task(_run_query_batch, jid, queries, req.format, req.delay)
    return {"job_id": jid, "total": len(queries)}


def _run_query_batch(jid, queries, fmt, delay):
    j = JOBS[jid]; j["status"] = "running"; results = []
    for i, query in enumerate(queries):
        j["message"] = f"Searching {i+1}/{len(queries)}: {query[:55]}\u2026"
        j["progress"] = int(i / len(queries) * 90)
        try:
            r = search_brand_website(query)
            r = _ensure_hd_images(r)
            if not r.get("GTIN"):
                r["GTIN"] = _extract_gtin(r)
            results.append(r); j["results"].append(r)
        except Exception as e:
            err = {"Product Name": query, "Error": str(e)}
            results.append(err); j["results"].append(err)
        j["completed"] = i + 1
        if i < len(queries) - 1:
            time.sleep(delay)
    try:
        fn = _save(results, fmt, jid); j["download_url"] = f"/download/{fn}"
    except Exception as e:
        j["error"] = str(e)
    j.update({"status": "done", "progress": 100,
               "message": f"Done \u2014 {len(results)} products scraped"})


# ── Category page scrape (async background job) ───────────────────────────────
@app.post("/scrape/category")
def scrape_category(req: CategoryScrapeRequest, background_tasks: BackgroundTasks):
    """
    Scrape all products from a category/listing page URL.
    Special handling for fouanistore.com and samsung.com.
    Generic link extraction + Playwright fallback for any other site.
    """
    if not req.url.strip().startswith(("http://", "https://")):
        raise HTTPException(400, "Invalid URL")
    if not CATEGORY_SCRAPER_AVAILABLE:
        raise HTTPException(503, "Category scraper module not available")
    jid = uuid.uuid4().hex[:8]
    _job(jid, req.max_products)
    background_tasks.add_task(
        _run_category, jid, req.url.strip(), req.max_products, req.format
    )
    return {"job_id": jid, "max_products": req.max_products}


def _run_category(jid, url, max_products, fmt):
    j = JOBS[jid]; j["status"] = "running"
    j["message"] = f"Starting category scrape: {url[:55]}…"

    def _progress(pct, msg):
        j["progress"] = int(pct)
        j["message"] = msg
        # Update completed count from progress percentage
        if j["total"] > 0:
            j["completed"] = max(j["completed"], int(pct / 100 * j["total"]))

    try:
        results = scrape_category_page(url, max_products=max_products,
                                       progress_cb=_progress)
        j["results"] = results
        j["completed"] = len(results)
        j["total"] = len(results)
        fn = _save(results, fmt, jid)
        j["download_url"] = f"/download/{fn}"
        j.update({"status": "done", "progress": 100,
                   "message": f"Done \u2014 {len(results)} products scraped"})
    except Exception as e:
        j.update({"status": "error", "error": str(e),
                   "message": f"Error: {str(e)[:120]}"})


@app.get("/download/{filename}")
def download(filename: str):
    safe=Path(filename).name; path=RESULTS_DIR/safe
    if not path.exists(): raise HTTPException(404,"File not found")
    mt=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if safe.endswith(".xlsx") else "text/csv")
    return FileResponse(path,media_type=mt,filename=safe,
                        headers={"Content-Disposition":f'attachment; filename="{safe}"'})
